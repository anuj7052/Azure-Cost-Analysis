"""
Write a BOQ back out in the Azure Pricing Calculator's own export layout.

The calculator's "Export" button produces a specific workbook, and procurement
teams read that shape without being told what it is. A BOQ exported in any other
shape has to be explained before it can be used, and it cannot be pasted next to
a real estimate for comparison.

The layout is reproduced from an actual export:

    A1  Microsoft Azure Estimate
    A2  Your Estimate
    A3  Service category | Service type | Custom name | Region | Description |
        Estimated monthly cost | Estimated upfront cost
    A4+ one row per line item
        Support / Licensing Program / Billing Account / Billing Profile / Total
        Disclaimer

Costs are what Azure actually billed, not list prices, and the disclaimer says
so — a reader who assumes calculator list prices from the familiar layout would
be reading the wrong number.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

CALCULATOR_URL = "https://azure.microsoft.com/pricing/calculator/"

HEADERS = [
    "Service category",
    "Service type",
    "Custom name",
    "Region",
    "Description",
    "Estimated monthly cost",
    "Estimated upfront cost",
]

# Column widths that make the description column readable without a manual
# resize, since the description is the longest field by a wide margin.
COLUMN_WIDTHS = [20, 26, 22, 18, 72, 22, 22]

# Symbols for the currencies this app is realistically used with. The disclaimer
# names the code either way, so an unknown currency degrades to the code alone
# rather than to a wrong symbol.
CURRENCY_NAMES = {
    "INR": ("India – Rupee", "₹"),
    "USD": ("US – Dollar", "$"),
    "EUR": ("Euro", "€"),
    "GBP": ("United Kingdom – Pound", "£"),
    "AUD": ("Australia – Dollar", "$"),
    "CAD": ("Canada – Dollar", "$"),
    "JPY": ("Japan – Yen", "¥"),
    "SGD": ("Singapore – Dollar", "$"),
    "AED": ("UAE – Dirham", "د.إ"),
}

# Service category is the calculator's top-level grouping ("Compute",
# "Networking"). Cost Management reports a service name instead, so the two are
# bridged by keyword. An unmatched service falls back to "Other" rather than
# being guessed into the wrong category, because a wrong category silently moves
# money between budget lines.
_CATEGORY_KEYWORDS = [
    ("Compute", ("virtual machine", "vm", "compute", "container", "kubernetes",
                 "app service", "function", "batch", "service fabric", "scale set")),
    ("Storage", ("storage", "disk", "blob", "file", "backup", "netapp", "data lake",
                 "archive", "recovery services")),
    ("Networking", ("network", "bandwidth", "load balancer", "gateway", "dns",
                    "front door", "traffic manager", "cdn", "expressroute",
                    "firewall", "public ip", "private link", "data transfer")),
    ("Databases", ("sql", "cosmos", "mysql", "postgres", "mariadb", "redis",
                   "database", "synapse")),
    ("Analytics", ("analytics", "databricks", "hdinsight", "stream analytics",
                   "event hub", "data factory", "purview")),
    ("AI + machine learning", ("cognitive", "openai", "machine learning",
                               "bot service", "search")),
    ("Security", ("key vault", "sentinel", "defender", "security")),
    ("Management and governance", ("monitor", "log analytics", "insights",
                                   "automation", "policy", "advisor")),
    ("Identity", ("active directory", "entra", "identity")),
    ("Integration", ("logic apps", "service bus", "api management",
                     "event grid", "notification")),
]


def category_for(service: str) -> str:
    """The calculator's service category for a Cost Management service name."""
    name = (service or "").lower()
    for category, keywords in _CATEGORY_KEYWORDS:
        if any(word in name for word in keywords):
            return category
    return "Other"


def describe(item: Dict[str, Any], currency: str) -> str:
    """
    The description cell: what a reader needs to price the line themselves.

    The calculator writes a sentence per line ("1 D2ls v6 (2 vCPUs, 4 GB RAM) x
    550 Hours (Pay as you go)…"), so this writes one too — quantity, spec, and
    the derived unit rate, which is the figure someone will want to check
    against a published price.
    """
    quantity = item.get("quantity") or 0
    spec = item.get("spec") or "Standard"
    unit = item.get("unit_monthly_cost")

    parts = [f"{quantity} x {spec}"]
    if unit is not None:
        parts.append(f"{unit:,.2f} {currency} per unit per month")

    priced = item.get("priced_quantity")
    if priced is not None and quantity and priced < quantity:
        # Silence here would present a partial total as a complete one.
        parts.append(f"{quantity - priced} of {quantity} had no billed cost reported")

    groups = item.get("resource_groups") or []
    if groups:
        shown = "; ".join(groups[:3])
        parts.append(f"Resource groups: {shown}" + (" …" if len(groups) > 3 else ""))

    examples = item.get("examples") or []
    if examples:
        parts.append("e.g. " + ", ".join(examples))

    return " — ".join(parts)


def _row(sheet, index: int, values: List[Any]) -> None:
    for column, value in enumerate(values, start=1):
        sheet.cell(row=index, column=column, value=value)


def build_estimate_workbook(
    boq: Dict[str, Any],
    *,
    title: str = "Your Estimate",
    source_note: Optional[str] = None,
) -> bytes:
    """
    Render a BOQ as an .xlsx in the pricing calculator's export layout.

    Returns bytes so the caller decides whether it is streamed, attached or
    written to disk.
    """
    currency = (boq.get("currency") or "USD").upper()
    label, symbol = CURRENCY_NAMES.get(currency, (currency, ""))
    items: List[Dict[str, Any]] = boq.get("items") or []

    book = Workbook()
    sheet = book.active
    sheet.title = "Estimate"

    bold = Font(bold=True)
    money = "#,##0.00"

    _row(sheet, 1, ["Microsoft Azure Estimate"])
    sheet["A1"].font = Font(bold=True, size=14)
    _row(sheet, 2, [title])
    sheet["A2"].font = bold

    _row(sheet, 3, HEADERS)
    for column in range(1, len(HEADERS) + 1):
        sheet.cell(row=3, column=column).font = bold

    line = 4
    for item in items:
        _row(sheet, line, [
            category_for(item.get("service", "")),
            item.get("service", "") or "Other",
            # The calculator leaves "Custom name" blank unless the user typed
            # one; the region-qualified spec is the closest honest equivalent
            # and keeps two lines of the same service distinguishable.
            item.get("spec", ""),
            item.get("region", ""),
            describe(item, currency),
            round(float(item.get("monthly_cost") or 0), 2),
            0,
        ])
        sheet.cell(row=line, column=6).number_format = money
        sheet.cell(row=line, column=7).number_format = money
        sheet.cell(row=line, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        line += 1

    # Support is a real calculator row and is always present, at zero unless a
    # support plan was quoted. Dropping it would make the export diverge from
    # the layout it is meant to match.
    _row(sheet, line, ["Support", "", "", "Support", "", 0, 0])
    line += 1

    _row(sheet, line, ["", "", "", "Licensing Program", "Microsoft Customer Agreement (MCA)"])
    line += 1
    _row(sheet, line, ["", "", "", "Billing Account", boq.get("billing_account", "")])
    line += 1
    _row(sheet, line, ["", "", "", "Billing Profile", boq.get("billing_profile", "")])
    line += 1

    total = round(float(boq.get("total_monthly") or 0), 2)
    _row(sheet, line, ["", "", "", "Total", "", total, 0])
    for column in (4, 6, 7):
        sheet.cell(row=line, column=column).font = bold
    sheet.cell(row=line, column=6).number_format = money
    sheet.cell(row=line, column=7).number_format = money
    line += 2

    _row(sheet, line, ["Disclaimer"])
    sheet.cell(row=line, column=1).font = bold
    line += 1

    prefix = f"{label} ({symbol}) {currency}" if symbol else f"{label} {currency}"
    _row(sheet, line, [
        f"All prices shown are in {prefix}. This is a summary estimate, not a quote. "
        f"For up to date pricing information please visit {CALCULATOR_URL}"
    ])
    line += 1

    # The provenance line is the one place this differs in substance from a real
    # calculator export, so it is stated plainly rather than left to inference.
    _row(sheet, line, [
        source_note
        or "Figures are the amounts Azure actually billed for these resources, "
           "not published list prices, so they already include any discount, "
           "reservation or negotiated rate."
    ])
    line += 1

    unpriced = int(boq.get("unpriced_count") or 0)
    if unpriced:
        _row(sheet, line, [
            f"{unpriced} resource(s) had no billed cost reported and contribute "
            f"nothing to the total. The total is therefore a floor, not a complete figure."
        ])
        line += 1

    # Built by hand rather than with strftime: the calculator writes an
    # unpadded US date, and the '%-d' form that produces one is not portable.
    now = datetime.now(timezone.utc)
    hour = now.hour % 12 or 12
    created = (
        f"{now.month}/{now.day}/{now.year} "
        f"{hour}:{now.minute:02d}:{now.second:02d} {'AM' if now.hour < 12 else 'PM'}"
    )
    _row(sheet, line, [f"This estimate was created at {created} UTC."])

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
